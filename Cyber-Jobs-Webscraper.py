'''
Luke Katsel
Cyber Jobs Webscraper created
for UArizona Cyber Security Clinic
12/14/24
'''
debug = True

import openpyxl     # Excel file reading and writing
import re           # Regular expressions for pattern matching
import requests     # Python Standard Library for url requests
import os           # Operating system interaction utilities
import csv
import pandas as pd
from jobspy import scrape_jobs
from bs4 import BeautifulSoup   # HTML and XML parsing tool

#path to excel file for url input and program output
excel_file_path = 'Cyber-Jobs.xlsx'

# regular expression pattern for urls
url_pattern = r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'

# open the Excel file
workbook = openpyxl.load_workbook(excel_file_path)

# initialize a list to store extracted URLs ands keywords
extracted_urls = []
keywords = []

# create a variable for the input sheet
# the second and third sheet (1 and 2) in the workbook will always be the input sheet
sheet_name = workbook.sheetnames[0]
excel_output = workbook[sheet_name]

sheet_name = workbook.sheetnames[1]
excel_filters = workbook[sheet_name]

sheet_name = workbook.sheetnames[2]
excel_urls = workbook[sheet_name]

# look in the first row of the filters worksheet for key words
for row in excel_filters.iter_rows(min_row=1, max_row=1, values_only=True): 
    #skip the first cell
    for cell_value in row[1:]:
        if cell_value is not None and isinstance(cell_value, str):    
            keywords.append(cell_value)
            
# for debugging 
if debug: 
    print(f"Debug: Keywords found: {keywords}")

#look in the secound row of the filters worksheet for location 
for row in excel_filters.iter_rows(min_row=2, max_row=2, values_only=True): 
    #skip the first cell
    for cell_value in row[1:]:
        if cell_value is not None and isinstance(cell_value, str):    
            location = cell_value

# This block is for the extraction of Urls from the excel sheet
'''
#iterate through each row to extract urls
for row in excel_urls.iter_rows(values_only=True):
    #iterate through each cell in row
    for cell_value in row:
        if cell_value is not None and isinstance(cell_value, str):
            # find URLs in the cell's text using the regular expression
            urls = re.findall(url_pattern, cell_value)
            # add URLs to extracted_urls list
            extracted_urls.extend(urls)

            
# url duplication sanitization 
unique_urls = list(set(extracted_urls))

# for debugging 
if debug: 
    print(f"Debug: unique urls: {unique_urls}")

# create list for filtered urls
filtered_urls = []

# insert filters into the URLs
for url in unique_urls:
    #if url == 'https://arizona.joinhandshake.com/explore':
        
    #if url == 'https://www.indeed.com/':
    
    #if url == 'https://www.linkedin.com/jobs/':
        
    #if url == 'https://www.cybersecurityjobs.com/':
        
    if url == 'https://www.ziprecruiter.com/':
        
        for word in keywords:
            #create the new urls
            new_url = f'https://www.ziprecruiter.com/jobs-search?search={word}&location={location}'
        
            filtered_urls.append(new_url)
        
    else: 
        # for debugging 
        if debug: 
            print(f"Debug: URL not recognized: {url}")        

# for debugging 
if debug: 
    print(f"Debug: filtered URLs: {filtered_urls}")
    
'''
# This is if we need headers or proxies// may need to change proxies
'''
headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36'
    }
proxy = {
    'http': 'http://35.247.237.139:31986',
    'http': 'http://52.26.114.229:1080'  
}

'''

# This block is for requesting the website from the url list
'''
# urls are ready for use
for url in filtered_urls:
    
    # retrieve a page from the website
    page = requests.get(url, headers=headers, proxies=proxy)
    
    # for debugging 
    if debug: 
        print(f"Debug: page: {page}")
        
    # convert the page into soup 
    soup = BeautifulSoup(page.text, 'html.parser')
    
    # for debugging 
    #if debug: 
        #print(f"Debug: soup: {soup}")
'''


for word in keywords:
    jobs = scrape_jobs(
        site_name=["indeed", "linkedin", "zip_recruiter", "glassdoor", "google"],
        search_term=f"{word}",
        google_search_term=f"{word}jobs near {location} since yesterday",
        location=f"{location}",
        results_wanted=20,
        hours_old=72,
        country_indeed='USA',
        
        linkedin_fetch_description=True # gets more info such as description, direct job url (slower)
        # proxies=["208.195.175.46:65095", "208.195.175.45:65095", "localhost"],
    )
    
    print(f"Found {len(jobs)} jobs for keyword '{word}'")
    #jobs.to_csv(f"{word}.csv", quoting=csv.QUOTE_NONNUMERIC, escapechar="\\", index=False) # to_excel      
    
    with pd.ExcelWriter('Cyber-Jobs.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        
        title_df = pd.DataFrame([[f"Results for keyword: {word}"]])  # Create a one-row DataFrame with the keyword as a title
        title_df.to_excel(writer, sheet_name="Output", startrow=writer.sheets["Output"].max_row, header=False, index=False)    
        
        jobs.to_excel(writer, sheet_name="Output", startrow=writer.sheets["Output"].max_row, index=False)
        

print('\n\nScript Complete')